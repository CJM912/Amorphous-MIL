import torch
from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
from sklearn.metrics import roc_auc_score

import numpy as np
from sklearn import metrics
import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import roc_curve
from sklearn.metrics import auc

probs = torch.load('probs-50.pth')
maxs = []
for i in range(15): # 测试集的数量（正例测试集和负例测试集）
    maxs.append(max(probs[i*1849:(i+1)*1849]))   # 1849的计算公式：{[(图片尺寸-tile尺寸)/步长] + 1}的平方 ; 步长 = tile尺寸/2 ;
targets = torch.load('test-50.lib')['targets']
#print(maxs)
preds = [1 if i>0.5 else 0 for i in maxs]
print(targets)
print(preds)
print(maxs)
print(confusion_matrix(targets, preds))
print(classification_report(targets, preds, digits=4))
print('AUC:',roc_auc_score(targets,maxs))

fpr, tpr, thresholds = roc_curve(targets, maxs,pos_label=1)
roc_auc = auc(fpr,tpr)
print(fpr)
print(tpr)
print(maxs)
from openpyxl import Workbook
from openpyxl import load_workbook

rs=Workbook()
rss=rs.active
fpr_li=list(fpr)
tpr_li=list(tpr)
for i in range(0,len(fpr)):
    rss.cell(row=i+1,column=1).value=fpr_li[i]

for j in range(0,len(tpr)):
    rss.cell(row=j+1,column=2).value=tpr_li[j]

rs.save('result_100.xlsx')

#plt.figure()
lw = 2
#plt.figure(figsize=(10,10))
plt.plot(fpr, tpr, color='darkorange',
         lw=lw, label='ROC curve (area = %0.2f)' % roc_auc) ###假正率为横坐标，真正率为纵坐标做曲线
plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="lower right")
#plt.show()

plt.savefig('./roc_50.png')



'''
plt.figure()
lw = 2
plt.plot(targets, maxs, color='darkorange',
         lw=lw, label='ROC curve (area = %0.2f)' % auc)
plt.plot([0, 1], [0, 1], color='navy', linestyle='--')
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="lower right")
plt.show()
'''
